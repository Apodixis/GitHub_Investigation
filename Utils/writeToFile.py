# Utils/writeToFile.py
import os, openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

search_info = {}

def write_to_excel(user_data: list, target: str, search_info: dict):
    """
    Inputs: List of dicts containing returned information and target name(s).
    Outputs: Excel file written to user's Downloads folder.
    Method: Uses Openpyxl to write data to an Excel file.
    """
    searchMode = search_info.get("search_mode")
    searchMethod = search_info.get("search_method")
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{searchMode}Search{searchMethod}"
    
    # Preserve column order: start with keys from first user, append any new keys found in others
    if user_data:
        all_keys = list(user_data[0].keys())
        for user in user_data[1:]:
            for k in user.keys():
                if k not in all_keys:
                    all_keys.append(k)
    else:
        all_keys = []
    
    # Write header
    for col, key in enumerate(all_keys, 1):
        ws.cell(row=1, column=col, value=key)
    
    # Write user data
    for row, user in enumerate(user_data, 2):
        for col, key in enumerate(all_keys, 1):
            val = user.get(key, "")
            # Convert sets/lists to comma-separated string for Excel
            if isinstance(val, (set, list)):
                val = ', '.join(str(item) for item in val)
            elif isinstance(val, dict):
                val = str(val)
            ws.cell(row=row, column=col, value=val)
    
    # Autosize columns
    for col in range(1, len(all_keys)+1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True
    
    # Build filename and path to Downloads
    date_str = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"{date_str}{searchMode}Search{searchMethod}_{target}.xlsx"
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    file_path = os.path.join(downloads_folder, filename)
    
    # Save results to workbook
    wb.save(file_path)
    return filename